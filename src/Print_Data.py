import pandas as pd
import json
import openpyxl
import datetime
from Bot_retrive_data import printDataAWS
from google_sheets import upload_excel

def clear_excel_data(file_path):
    """
    Clears all the data from an Excel file.

    Args:
        file_path: The path to the Excel file.

    Returns:
        None
    """
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)

    # Iterate over all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Clear the content of each cell in the sheet
        sheet.delete_rows(1, sheet.max_row)

    # Save the changes to the Excel file
    workbook.save(file_path)

def adjust_column_width(excel_file):
    """
    Adjusts the column width in the Excel file to accommodate the data entered.

    Args:
        excel_file: The path to the Excel file.

    Returns:
        None
    """
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file)

    # Iterate over all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Adjust the column width for each column
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the changes to the Excel file
    workbook.save(excel_file)

def print_json_to_excel():
    # Load the JSON data
    json_data = printDataAWS()
    excel_file = '/Users/gaurav.gupta2/Documents/program/botbuilder-samples/samples/python/05.multi-turn-prompt/Spot_Price_Comparision.xlsx'
    clear_excel_data(excel_file)
    data = json.loads(json_data)
    timestamp = int(datetime.datetime.now().timestamp())
    desired_columns = ['InstanceType', 'vCPU', 'Memory','Ratio','Region','Price (Per Month)','Category']  # Replace with your desired column names

    # Filter the JSON data to include only the desired columns
    filtered_data = [{column: item.get(column) for column in desired_columns} for item in data]

    # Convert the filtered data to a DataFrame
    df = pd.DataFrame(filtered_data)

    # Write the DataFrame to an Excel file
    df.to_excel(excel_file, index=False)
    adjust_column_width(excel_file)
    return upload_excel(excel_file)

if __name__ == '__main__':
    print(print_json_to_excel())
